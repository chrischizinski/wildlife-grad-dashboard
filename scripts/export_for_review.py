#!/usr/bin/env python3
"""
Export all positions to an Excel spreadsheet for comprehensive review and correction.

This script creates a detailed Excel file with all position data, allowing for
batch review and corrections of ML classifications. The spreadsheet includes
dropdown menus and conditional formatting to make review efficient.

Usage:
    python scripts/export_for_review.py
    python scripts/export_for_review.py --output my_review.xlsx
    python scripts/export_for_review.py --unverified-only
"""

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path
from typing import Dict, List

try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.formatting.rule import CellIsRule
except ImportError:
    print("❌ Required packages not installed. Please run:")
    print("   pip install pandas openpyxl")
    sys.exit(1)


class ReviewSpreadsheetExporter:
    """Export positions to Excel for comprehensive review."""

    def __init__(self):
        self.data_file = Path("data/processed/verified_graduate_assistantships.json")
        self.positions: List[Dict] = []
        
        # Available options for corrections
        self.disciplines = [
            "Wildlife & Natural Resources",
            "Natural Resource Management", 
            "Environmental Science",
            "Fisheries & Aquatic Sciences",
            "Conservation Biology",
            "Unknown",
            "Other",
        ]
        
        self.position_types = [
            "Graduate",
            "Professional", 
            "Technician",
            "Exclude",
        ]
        
        self.graduate_options = ["Yes", "No"]
        self.big10_options = ["Yes", "No", "Unknown"]
        self.review_status = ["✅ Correct", "🔧 Needs Correction", "❌ Exclude", "⏭️ Skip"]

    def load_data(self) -> None:
        """Load position data."""
        if not self.data_file.exists():
            print(f"❌ Data file not found: {self.data_file}")
            sys.exit(1)

        with open(self.data_file, "r") as f:
            self.positions = json.load(f)

        print(f"📊 Loaded {len(self.positions)} positions for review")

    def prepare_dataframe(self, unverified_only: bool = False) -> pd.DataFrame:
        """Prepare DataFrame with all position data and review columns."""
        
        # Filter to unverified only if requested
        positions = self.positions
        if unverified_only:
            positions = [p for p in positions if not p.get("human_verified", False)]
            print(f"📋 Filtered to {len(positions)} unverified positions")

        # Prepare data for DataFrame
        data = []
        for pos in positions:
            # Basic position info
            row = {
                'Position_ID': pos.get('position_id', ''),
                'Title': pos.get('title', ''),
                'Organization': pos.get('organization', ''),
                'Location': pos.get('location', ''),
                'Salary': pos.get('salary', ''),
                'Starting_Date': pos.get('starting_date', ''),
                'Scraped_Date': pos.get('scraped_at', '')[:10] if pos.get('scraped_at') else '',
                
                # Current ML Classifications
                'ML_Graduate_Position': 'Yes' if pos.get('is_graduate_position') else 'No',
                'ML_Graduate_Confidence': round(pos.get('grad_confidence', 0), 3),
                'ML_Position_Type': pos.get('position_type', ''),
                'ML_Discipline': pos.get('discipline', ''),
                'ML_Discipline_Confidence': round(pos.get('discipline_confidence', 0), 3),
                'ML_Keywords': ', '.join(pos.get('discipline_keywords', [])[:5]),
                'ML_University_Name': pos.get('university_name', ''),
                'ML_Big10_University': 'Yes' if pos.get('is_big10_university') else 'No',
                
                # Verification status
                'Human_Verified': 'Yes' if pos.get('human_verified', False) else 'No',
                
                # Review columns (empty for user to fill)
                'Review_Status': '',
                'Corrected_Graduate_Position': '',
                'Corrected_Position_Type': '',
                'Corrected_Discipline': '',
                'Corrected_University_Name': '',
                'Corrected_Big10_University': '',
                'Review_Notes': '',
                'Reviewer_Name': '',
                
                # Description (truncated for display)
                'Description_Preview': (pos.get('description', '')[:200] + '...' 
                                      if len(pos.get('description', '')) > 200 
                                      else pos.get('description', '')),
                
                # Full description (hidden column for reference)
                'Full_Description': pos.get('description', ''),
            }
            data.append(row)

        df = pd.DataFrame(data)
        
        # Sort by confidence (lowest first) to prioritize review
        df = df.sort_values(['ML_Graduate_Confidence', 'ML_Discipline_Confidence'])
        
        return df

    def create_excel_file(self, df: pd.DataFrame, output_file: str) -> None:
        """Create formatted Excel file with dropdowns and styling."""
        print(f"📝 Creating Excel file: {output_file}")
        
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Position Review"
        
        # Add data to worksheet
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # Style the header row
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Set column widths
        column_widths = {
            'A': 15,  # Position_ID
            'B': 50,  # Title
            'C': 40,  # Organization
            'D': 25,  # Location
            'E': 15,  # Salary
            'F': 15,  # Starting_Date
            'G': 15,  # Scraped_Date
            'H': 20,  # ML_Graduate_Position
            'I': 20,  # ML_Graduate_Confidence
            'J': 20,  # ML_Position_Type
            'K': 30,  # ML_Discipline
            'L': 20,  # ML_Discipline_Confidence
            'M': 40,  # ML_Keywords
            'N': 30,  # ML_University_Name
            'O': 20,  # ML_Big10_University
            'P': 20,  # Human_Verified
            'Q': 20,  # Review_Status
            'R': 25,  # Corrected_Graduate_Position
            'S': 25,  # Corrected_Position_Type
            'T': 30,  # Corrected_Discipline
            'U': 30,  # Corrected_University_Name
            'V': 25,  # Corrected_Big10_University
            'W': 50,  # Review_Notes
            'X': 20,  # Reviewer_Name
            'Y': 60,  # Description_Preview
            'Z': 20,  # Full_Description (hidden)
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Hide the full description column
        ws.column_dimensions['Z'].hidden = True
        
        # Add data validation (dropdowns) for review columns
        last_row = len(df) + 1
        
        # Review Status dropdown
        review_status_dv = DataValidation(
            type="list",
            formula1=f'"{",".join(self.review_status)}"',
            showDropDown=True
        )
        ws.add_data_validation(review_status_dv)
        review_status_dv.add(f'Q2:Q{last_row}')
        
        # Graduate Position dropdown
        graduate_dv = DataValidation(
            type="list", 
            formula1=f'"{",".join(self.graduate_options)}"',
            showDropDown=True
        )
        ws.add_data_validation(graduate_dv)
        graduate_dv.add(f'R2:R{last_row}')
        
        # Position Type dropdown
        position_type_dv = DataValidation(
            type="list",
            formula1=f'"{",".join(self.position_types)}"',
            showDropDown=True
        )
        ws.add_data_validation(position_type_dv)
        position_type_dv.add(f'S2:S{last_row}')
        
        # Discipline dropdown
        discipline_dv = DataValidation(
            type="list",
            formula1=f'"{",".join(self.disciplines)}"',
            showDropDown=True
        )
        ws.add_data_validation(discipline_dv)
        discipline_dv.add(f'T2:T{last_row}')
        
        # Big 10 dropdown
        big10_dv = DataValidation(
            type="list",
            formula1=f'"{",".join(self.big10_options)}"',
            showDropDown=True
        )
        ws.add_data_validation(big10_dv)
        big10_dv.add(f'V2:V{last_row}')
        
        # Add conditional formatting for confidence levels
        # Low confidence (< 0.7) = Red
        low_conf_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        low_conf_rule = CellIsRule(operator="lessThan", formula=["0.7"], fill=low_conf_fill)
        ws.conditional_formatting.add(f'I2:I{last_row}', low_conf_rule)
        ws.conditional_formatting.add(f'L2:L{last_row}', low_conf_rule)
        
        # Medium confidence (0.7-0.9) = Yellow
        med_conf_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        med_conf_rule = CellIsRule(operator="between", formula=["0.7", "0.9"], fill=med_conf_fill)
        ws.conditional_formatting.add(f'I2:I{last_row}', med_conf_rule)
        ws.conditional_formatting.add(f'L2:L{last_row}', med_conf_rule)
        
        # Highlight "Unknown" disciplines
        unknown_fill = PatternFill(start_color="FFE4B5", end_color="FFE4B5", fill_type="solid")
        unknown_rule = CellIsRule(operator="equal", formula=['"Unknown"'], fill=unknown_fill)
        ws.conditional_formatting.add(f'K2:K{last_row}', unknown_rule)
        
        # Freeze panes to keep headers visible
        ws.freeze_panes = "A2"
        
        # Add instructions sheet
        self.add_instructions_sheet(wb)
        
        # Save the file
        wb.save(output_file)
        print(f"✅ Excel file created: {output_file}")
        print(f"📝 {len(df)} positions ready for review")

    def add_instructions_sheet(self, wb: Workbook) -> None:
        """Add instructions sheet to the workbook."""
        ws = wb.create_sheet("Instructions", 0)  # Insert as first sheet
        
        instructions = [
            ["Wildlife Graduate Assistantships - Position Review Instructions", ""],
            ["", ""],
            ["OVERVIEW:", ""],
            ["This spreadsheet contains all position data with ML classifications", ""],
            ["for your review. Use the dropdown menus to make corrections.", ""],
            ["", ""],
            ["COLUMN GUIDE:", ""],
            ["• Columns A-P: Position data and current ML classifications", ""],
            ["• Columns Q-X: Your review and corrections", ""],
            ["• Column Y: Description preview for reference", ""],
            ["", ""],
            ["HOW TO REVIEW:", ""],
            ["1. Review Status: Mark as ✅ Correct, 🔧 Needs Correction, ❌ Exclude, or ⏭️ Skip", ""],
            ["2. If 'Needs Correction', fill in the correction columns", ""],
            ["3. Add notes in the Review_Notes column", ""],
            ["4. Add your name in Reviewer_Name column", ""],
            ["", ""],
            ["PRIORITY INDICATORS:", ""],
            ["• Red highlighting = Low confidence ML prediction (< 0.7)", ""],
            ["• Yellow highlighting = Medium confidence (0.7-0.9)", ""],
            ["• Orange highlighting = 'Unknown' discipline (needs attention)", ""],
            ["", ""],
            ["AVAILABLE OPTIONS:", ""],
            ["Graduate Position: Yes, No", ""],
            ["Position Types: Graduate, Professional, Technician, Exclude", ""],
            ["Disciplines: Wildlife & Natural Resources, Natural Resource Management,", ""],
            ["             Environmental Science, Fisheries & Aquatic Sciences,", ""],
            ["             Conservation Biology, Unknown, Other", ""],
            ["Big 10 Universities: Yes, No, Unknown", ""],
            ["", ""],
            ["TIPS:", ""],
            ["• Focus on red/yellow highlighted rows first (low confidence)", ""],
            ["• Look for 'Unknown' disciplines that can be classified", ""],
            ["• Use the description preview to understand the position", ""],
            ["• Save frequently to avoid losing work", ""],
            ["", ""],
            ["When complete, save this file and run:", ""],
            ["python scripts/import_review_corrections.py your_file.xlsx", ""],
        ]
        
        for row_idx, (instruction, note) in enumerate(instructions, 1):
            ws[f'A{row_idx}'] = instruction
            ws[f'B{row_idx}'] = note
        
        # Style the instructions
        ws['A1'].font = Font(bold=True, size=16)
        
        # Set column widths
        ws.column_dimensions['A'].width = 60
        ws.column_dimensions['B'].width = 40
        
        # Add borders and styling to make it readable
        for row in range(1, len(instructions) + 1):
            if ws[f'A{row}'].value and ws[f'A{row}'].value.endswith(':'):
                ws[f'A{row}'].font = Font(bold=True)

    def create_summary_stats(self, df: pd.DataFrame) -> None:
        """Print summary statistics about the data."""
        print(f"\n📈 Review Summary Statistics:")
        print(f"   Total positions: {len(df)}")
        
        # Verification status
        verified_count = len(df[df['Human_Verified'] == 'Yes'])
        unverified_count = len(df) - verified_count
        print(f"   Already verified: {verified_count}")
        print(f"   Need review: {unverified_count}")
        
        # Confidence distribution
        low_grad_conf = len(df[df['ML_Graduate_Confidence'] < 0.7])
        medium_grad_conf = len(df[(df['ML_Graduate_Confidence'] >= 0.7) & (df['ML_Graduate_Confidence'] < 0.9)])
        high_grad_conf = len(df[df['ML_Graduate_Confidence'] >= 0.9])
        
        print(f"\n🎯 Graduate Classification Confidence:")
        print(f"   Low (< 0.7): {low_grad_conf} positions")
        print(f"   Medium (0.7-0.9): {medium_grad_conf} positions") 
        print(f"   High (≥ 0.9): {high_grad_conf} positions")
        
        # Discipline distribution
        unknown_disciplines = len(df[df['ML_Discipline'] == 'Unknown'])
        total_disciplines = df['ML_Discipline'].nunique()
        
        print(f"\n🔬 Discipline Classification:")
        print(f"   Unknown disciplines: {unknown_disciplines} positions")
        print(f"   Total unique disciplines: {total_disciplines}")
        
        # Top disciplines
        top_disciplines = df['ML_Discipline'].value_counts().head(5)
        print(f"\n   Top 5 disciplines:")
        for discipline, count in top_disciplines.items():
            print(f"     {discipline}: {count}")


def main():
    parser = argparse.ArgumentParser(
        description="Export positions to Excel spreadsheet for comprehensive review"
    )
    
    parser.add_argument(
        "--output", 
        type=str, 
        default=f"position_review_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        help="Output Excel file name"
    )
    
    parser.add_argument(
        "--unverified-only",
        action="store_true", 
        help="Export only unverified positions (exclude already human-verified ones)"
    )
    
    args = parser.parse_args()
    
    print("🚀 Wildlife Graduate Assistantships - Review Export")
    print("=" * 60)
    
    exporter = ReviewSpreadsheetExporter()
    exporter.load_data()
    
    # Prepare DataFrame
    df = exporter.prepare_dataframe(unverified_only=args.unverified_only)
    
    if df.empty:
        print("❌ No positions to export")
        return
    
    # Show summary stats
    exporter.create_summary_stats(df)
    
    # Create Excel file
    exporter.create_excel_file(df, args.output)
    
    print(f"\n✅ Review spreadsheet ready!")
    print(f"📄 File: {args.output}")
    print(f"📊 Positions to review: {len(df)}")
    print(f"\n💡 Next steps:")
    print(f"   1. Open the Excel file and review positions")
    print(f"   2. Focus on red/yellow highlighted rows (low confidence)")
    print(f"   3. When done, run: python scripts/import_review_corrections.py {args.output}")


if __name__ == "__main__":
    main()